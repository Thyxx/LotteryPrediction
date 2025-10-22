from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import datetime
from typing import Dict, Iterable, List, Optional, Tuple

import requests
from dateutil import parser
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import EuroMillionsDraw, LotoDraw

from openpyxl import load_workbook


LOTTO_URL = (
    "https://www.loterie-nationale.be/content/dam/opp/draw-games/lotto/"
    "brand-assets/documents/fr/statistiques-lotto-08-25.xlsx"
)
EUROMILLIONS_URL = "https://media.fdj.fr/static/csv/euromillions.csv"


class FetchError(RuntimeError):
    """Raised when lottery data cannot be fetched."""


def _download_csv(url: str) -> str:
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    response.encoding = "utf-8"
    return response.text


def _download_binary(url: str) -> bytes:
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return response.content


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = normalized.lower()
    translation = str.maketrans(
        {
            " ": "_",
            "-": "_",
            ".": "",
            ",": "",
            ";": "",
            ":": "",
            "'": "",
            "’": "",
            "(": "",
            ")": "",
            "/": "_",
            "\\": "_",
            "°": "",
        }
    )
    normalized = normalized.translate(translation)
    normalized = re.sub(r"([a-z])([0-9])", r"\1_\2", normalized)
    normalized = re.sub(r"([0-9])([a-z])", r"\1_\2", normalized)
    normalized = re.sub(r"__+", "_", normalized)
    return normalized.strip("_")


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _looks_like_lotto_header(headers: List[str]) -> bool:
    if not headers:
        return False
    has_date = any("date" in header for header in headers)
    main_candidates = sum(
        1
        for header in headers
        if header.startswith(("boule", "numero", "num", "n_")) and "tirage" not in header
    )
    has_extra = any(
        keyword in header
        for keyword in ("chance", "bonus", "complementaire")
        for header in headers
    )
    return has_date and main_candidates >= 5 and has_extra


def _prepare_lotto_rows(xlsx_content: bytes) -> Iterable[Dict[str, str]]:
    workbook = load_workbook(io.BytesIO(xlsx_content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    header_index: Optional[int] = None
    header_map: Dict[int, str] = {}

    for index, row in enumerate(rows):
        normalized_row = [
            _normalize_header(str(cell)) if cell not in (None, "") else ""
            for cell in row
        ]
        if not any(normalized_row):
            continue
        if _looks_like_lotto_header(normalized_row):
            header_index = index
            header_map = {idx: header for idx, header in enumerate(normalized_row) if header}
            break

    if header_index is None:
        raise FetchError("Impossible de détecter l'en-tête du fichier Lotto.")

    for row in rows[header_index + 1 :]:
        row_data: Dict[str, str] = {}
        for idx, cell in enumerate(row):
            key = header_map.get(idx)
            if not key:
                continue
            value = _format_cell(cell)
            if value == "":
                continue
            row_data[key] = value
        if not row_data:
            continue
        yield row_data


def _prepare_reader(csv_content: str) -> Iterable[Dict[str, str]]:
    handle = io.StringIO(csv_content)
    sample = handle.read(4096)
    handle.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        reader = csv.DictReader(handle, dialect=dialect)
    except csv.Error:
        reader = csv.DictReader(handle, delimiter=';')
    # Normalize keys once here for easier handling downstream
    for row in reader:
        yield {key.strip().lower(): value.strip() for key, value in row.items() if key}


def _extract_numbers(row: Dict[str, str], prefix: str) -> List[int]:
    extracted: List[Tuple[int, int]] = []
    for key, value in row.items():
        if (
            not value
            or "tirage" in key
            or "chance" in key
            or "bonus" in key
            or "complementaire" in key
        ):
            continue
        if key.startswith(prefix):
            suffix = key.replace(prefix, "").lstrip("_")
            match = re.match(r"(\d+)", suffix)
            if not match:
                continue
            parsed = _parse_int(value)
            if parsed is None:
                continue
            extracted.append((int(match.group(1)), parsed))
    extracted.sort(key=lambda item: item[0])
    return [value for _, value in extracted]


def _parse_int(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None
    value = value.strip()
    if not value:
        return None
    try:
        return int(value)
    except ValueError as exc:  # pragma: no cover - defensive
        raise FetchError(f"Valeur numérique invalide: {value}") from exc


def _parse_date(row: Dict[str, str]) -> datetime:
    for key in ("date_de_tirage", "date_du_tirage", "date", "drawdate"):
        if key in row and row[key]:
            try:
                return parser.parse(row[key], dayfirst=True)
            except (ValueError, parser.ParserError) as exc:  # pragma: no cover
                raise FetchError(f"Date invalide: {row[key]}") from exc
    raise FetchError("Champ de date introuvable dans la ligne")


def _parse_draw_number(row: Dict[str, str]) -> Optional[int]:
    for key in (
        "numero_de_tirage",
        "numero_du_tirage",
        "num_tirage",
        "n_tirage",
        "tirage",
        "drawnumber",
    ):
        if key in row:
            return _parse_int(row[key])
    return None


def update_loto_draws(session: Session) -> int:
    """Download the latest Loto draws and merge them into the database."""
    xlsx_content = _download_binary(LOTTO_URL)
    reader = _prepare_lotto_rows(xlsx_content)

    inserted = 0
    for row in reader:
        draw_date = _parse_date(row).date()
        draw_number = _parse_draw_number(row)
        number_prefixes = ("boule_", "boule", "numero_", "num_", "n_")
        numbers: List[int] = []
        for prefix in number_prefixes:
            if numbers:
                break
            numbers = _extract_numbers(row, prefix)
        if not numbers:
            # fallback: gather numeric columns by position
            extracted: List[int] = []
            for key, value in row.items():
                if any(term in key for term in ("date", "tirage", "chance", "bonus", "complementaire")):
                    continue
                parsed = _parse_int(value)
                if parsed is not None:
                    extracted.append(parsed)
            numbers = sorted(set(extracted))
        chance_number: Optional[int] = None
        for key, value in row.items():
            if any(term in key for term in ("chance", "bonus", "complementaire")):
                chance_number = _parse_int(value)
                if chance_number is not None:
                    break
        if len(numbers) < 5 or chance_number is None:
            # Skip malformed entries
            continue

        numbers = sorted(set(numbers))
        # Some datasets may include duplicates; ensure exactly 5 numbers
        if len(numbers) != 5:
            continue

        stmt = select(LotoDraw).where(
            LotoDraw.draw_date == draw_date,
            LotoDraw.draw_number == draw_number,
        )
        existing = session.execute(stmt).scalar_one_or_none()
        if existing:
            continue

        draw = LotoDraw(
            draw_date=draw_date,
            draw_number=draw_number,
            main_numbers=",".join(str(num) for num in numbers),
            chance_number=chance_number,
        )
        session.add(draw)
        inserted += 1

    if inserted:
        session.commit()
    return inserted


def update_euromillions_draws(session: Session) -> int:
    """Download the latest EuroMillions draws and merge them into the database."""
    csv_content = _download_csv(EUROMILLIONS_URL)
    reader = _prepare_reader(csv_content)

    inserted = 0
    for row in reader:
        draw_date = _parse_date(row).date()
        draw_number = _parse_draw_number(row)
        numbers = _extract_numbers(row, "boule_")
        stars = _extract_numbers(row, "etoile_")
        if len(numbers) < 5 or len(stars) < 2:
            continue

        numbers = sorted(set(numbers))
        stars = sorted(set(stars))
        if len(numbers) != 5 or len(stars) != 2:
            continue

        stmt = select(EuroMillionsDraw).where(
            EuroMillionsDraw.draw_date == draw_date,
            EuroMillionsDraw.draw_number == draw_number,
        )
        existing = session.execute(stmt).scalar_one_or_none()
        if existing:
            continue

        draw = EuroMillionsDraw(
            draw_date=draw_date,
            draw_number=draw_number,
            main_numbers=",".join(str(num) for num in numbers),
            star_numbers=",".join(str(num) for num in stars),
        )
        session.add(draw)
        inserted += 1

    if inserted:
        session.commit()
    return inserted


def update_all_draws(session: Session) -> Dict[str, int]:
    """Update both Loto and EuroMillions draws, returning inserted counts."""
    results = {
        "loto": update_loto_draws(session),
        "euromillions": update_euromillions_draws(session),
    }
    return results
